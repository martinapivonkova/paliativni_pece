import openpyxl
import sys
import os

slozka_vstup = sys.argv[1]      #'zemreli2011-2017'
zemreli = sys.argv[2]           #'vstupy/zemreli2011-2017.csv'

for soubor in os.listdir(slozka_vstup):
    cesta = os.path.join(slozka_vstup, soubor)
    workbook = openpyxl.load_workbook(cesta)
    sheet = workbook.active
        
    c2 = sheet['C2'].value.split(' ')
    pohlavi = c2[0]
    rok = c2[-1]
    vekove_kategorie = [sheet.cell(row=3, column=c).value for c in range(6,25)]
    
    file = open(zemreli,'a', encoding='utf-8')
    row = 7
    while sheet.cell(row=row,column=2).value != None:
        nazev = sheet.cell(row=row, column=2)
        if not nazev.font.bold:
            mkn = sheet.cell(row=row, column=1).value 
            pocet_zemrelych = sum(int(str(sheet.cell(row=row, column=col).value).replace("-","0")) for col in range(4,6))
            hodnoty = [str(rok), pohlavi, mkn, nazev.value.replace("\n","").replace(",","-"), '0-4', str(pocet_zemrelych).replace("-","0")]
            file.write(','.join(hodnoty)+'\n')
            for col in range(6,25):
                pocet_zemrelych = sheet.cell(row=row, column=col).value
                hodnoty = [str(rok), pohlavi, mkn, nazev.value.replace("\n","").replace(",","-"), str(vekove_kategorie[col-6]).replace(" ","").replace("5-9","05-9"), str(pocet_zemrelych).replace("-","0")]
                file.write(','.join(hodnoty)+'\n')
        row = row+1   
    file.close()
        